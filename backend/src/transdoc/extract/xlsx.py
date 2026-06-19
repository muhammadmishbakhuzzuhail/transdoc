# SPDX-License-Identifier: AGPL-3.0-only
# Copyright (C) 2026 Muhammad Mishbakhuz Zuhail
"""Excel (.xlsx) extraction via openpyxl.

Two extraction shapes, picked from the output format:

* xlsx/same-as-source round-trip -> one IR block per non-empty string cell, id =
  ``{sheet}!{coordinate}`` (e.g. ``Sheet1!B3``). The renderer reopens the workbook and
  writes translations back to the same coordinates, keeping styles, column widths, merged
  ranges, and formulas intact.
* any other target (md/docx/pdf/plain) -> one TABLE block per sheet, so the grid renders
  as a real table instead of a flat one-cell-per-line list.

Numbers, dates, and formulas are left as-is — only text cells are translated.
"""

from __future__ import annotations

from ..config import Config, OutputFormat
from ..ir import Block, BlockType, Cell, Confidence, Document, Table
from .base import reflow_order

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def extract(path: str, cfg: Config) -> Document:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        raise ValueError(f"unreadable or corrupt XLSX: {e}") from e
    out = Document(source_path=path, mime=_MIME)

    # Round-trip back to a spreadsheet keeps the per-cell, coordinate-keyed shape so the
    # renderer can write each translation back in place. Cross-format output gets a grid.
    roundtrip = cfg.output_format in (OutputFormat.XLSX, OutputFormat.SAME)

    page = 0
    for ws in wb.worksheets:
        if roundtrip:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.strip() and not v.startswith("="):
                        out.blocks.append(Block(
                            id=f"{ws.title}!{cell.coordinate}", type=BlockType.PARAGRAPH,
                            page=page, text=v, confidence=Confidence(source="digital")))
        else:
            rows: list[list[Cell]] = []
            for row in ws.iter_rows():
                cells = [Cell(text="" if c.value is None else str(c.value)) for c in row]
                if any(c.text.strip() for c in cells):
                    rows.append(cells)
            if rows:
                out.blocks.append(Block(
                    id=f"{ws.title}!table", type=BlockType.TABLE, page=page,
                    table=Table(rows=rows), confidence=Confidence(source="digital")))
        page += 1

    reflow_order(out)
    return out
